#!/usr/bin/env python3
"""
CEAGESP Balanço Estatístico - Processador v6.0
Reescrito do zero usando openpyxl com dados reais dos arquivos 2025.

FILOSOFIA:
- Copia o template (mês anterior) e atualiza apenas as células de dados
- Lê valores existentes no template para acumular YTD corretamente
- CSV fornece apenas os dados do mês atual (ETSP e/ou Interior)
"""

import csv as csvmod
import io
import re
from collections import defaultdict
from copy import copy
import openpyxl
from openpyxl import load_workbook

# ── CONSTANTES ────────────────────────────────────────────────────────────────
GRUPOS = ['FRUTAS', 'LEGUMES', 'VERDURAS', 'DIVERSOS', 'FLORES', 'PESCADOS', 'ORGÂNICOS']
GRUPOS_INT = ['FRUTAS', 'LEGUMES', 'VERDURAS', 'DIVERSOS', 'FLORES', 'PESCADOS']  # interior sem ORGÂNICOS

GRUPO_MAP = {
    'FRUTAS': 'FRUTAS', 'LEGUMES': 'LEGUMES', 'VERDURAS': 'VERDURAS',
    'DIVERSOS': 'DIVERSOS', 'FLORES': 'FLORES', 'PESCADOS': 'PESCADOS',
    'ORGÂNICOS': 'ORGÂNICOS', 'ORGANICOS': 'ORGÂNICOS',
    'PESCADO': 'PESCADOS',
    'PRODUTOS LIGADOS À FLORICULTURA': 'FLORES',
}

MONTHS_PT = ['JANEIRO','FEVEREIRO','MARÇO','ABRIL','MAIO','JUNHO',
             'JULHO','AGOSTO','SETEMBRO','OUTUBRO','NOVEMBRO','DEZEMBRO']
MONTHS_SHORT = ['JAN','FEV','MAR','ABR','MAI','JUN','JUL','AGO','SET','OUT','NOV','DEZ']

# Mapeamento: código da sheet → nome da unidade no GERAL
SHEET_TO_UNIT = {
    'Ceara': 'ARAÇATUBA', 'Cearb': 'ARARAQUARA', 'Cebau': 'BAURU',
    'Cefra': 'FRANCA', 'Cemar': 'MARÍLIA', 'Cepir': 'PIRACICABA',
    'Cepre': 'PRESIDENTE PRUDENTE', 'Cerib': 'RIBEIRÃO PRETO',
    'Cesjc': 'SÃO JOSÉ DOS CAMPOS', 'Cesjr': 'SÃO JOSÉ DO RIO PRETO',
    'Cesor': 'SOROCABA', 'Cegua': 'GUARATINGUETÁ',
}

# CSV UNIDADE → código sheet
CSV_TO_SHEET = {
    'ENTREPOSTOS DA CAPITAL': 'ETSP',
    'ENTREPOSTO DA CAPITAL': 'ETSP',
    'CEASA DE SAO JOSE DO RIO PRETO': 'Cesjr',
    'CEASA DE RIBEIRAO PRETO': 'Cerib',
    'CEASA DE BAURU': 'Cebau',
    'CEASA DE SOROCABA': 'Cesor',
    'CEASA DE SAO JOSE DOS CAMPOS': 'Cesjc',
    'CEASA DE PRESIDENTE PRUDENTE': 'Cepre',
    'CEASA DE ARACATUBA': 'Ceara',
    'CEASA DE ARARAQUARA': 'Cearb',
    'CEASA DE FRANCA': 'Cefra',
    'CEASA DE MARILIA': 'Cemar',
    'CEASA DE PIRACICABA': 'Cepir',
    'CEASA DE GUARATINGUETA': 'Cegua',
}

# ── CSV PARSING ───────────────────────────────────────────────────────────────
def pf(v):
    try: return float(str(v or 0).replace(',', '.'))
    except: return 0.0

def parse_csv_content(content):
    """Parse CSV text (latin1 decoded) → list of dicts"""
    rows = []
    lines = content.replace('\r\n', '\n').replace('\r', '\n').split('\n')
    if not lines:
        return rows
    hdr = [h.strip().strip('"') for h in lines[0].split(';')]
    for line in lines[1:]:
        if not line.strip():
            continue
        vals = [v.strip().strip('"') for v in line.split(';')]
        rows.append({hdr[i]: vals[i] if i < len(vals) else '' for i in range(len(hdr))})
    return rows

def detect_unit(rows):
    """Detect which unit the CSV belongs to"""
    for row in rows[:20]:
        u = (row.get('UNIDADE', '') or '').strip().upper()
        c = CSV_TO_SHEET.get(u)
        if c:
            return c
    return None

def agg_by_grupo(rows):
    """Aggregate rows → {grupo: {ton, fin}}"""
    r = {g: {'ton': 0.0, 'fin': 0.0} for g in GRUPOS}
    for row in rows:
        g = GRUPO_MAP.get((row.get('GRUPO_PRODUTO', '') or '').strip().upper())
        if g:
            r[g]['ton'] += pf(row.get('PESO_TONELADA', 0))
            r[g]['fin'] += pf(row.get('VL_FINANCEIRO', 0))
    return r

def agg_by_produto(rows, grupo_filter=None):
    r = defaultdict(lambda: {'ton': 0.0, 'fin': 0.0})
    for row in rows:
        if grupo_filter:
            g = GRUPO_MAP.get((row.get('GRUPO_PRODUTO', '') or '').strip().upper())
            if g != grupo_filter:
                continue
        p = (row.get('DS_PRODUTO', '') or '').strip().upper()
        if p:
            r[p]['ton'] += pf(row.get('PESO_TONELADA', 0))
            r[p]['fin'] += pf(row.get('VL_FINANCEIRO', 0))
    return dict(r)

def agg_by_pais(rows):
    r = defaultdict(lambda: {'ton': 0.0, 'fin': 0.0})
    for row in rows:
        p = (row.get('PAIS', '') or 'BRASIL').strip().upper() or 'BRASIL'
        r[p]['ton'] += pf(row.get('PESO_TONELADA', 0))
        r[p]['fin'] += pf(row.get('VL_FINANCEIRO', 0))
    return dict(r)

def agg_by_uf(rows):
    r = defaultdict(lambda: {'ton': 0.0, 'fin': 0.0})
    for row in rows:
        uf = (row.get('DESCRICAO_ESTADO', '') or row.get('UF', '')).strip().upper() or 'NÃO INFORMADO'
        r[uf]['ton'] += pf(row.get('PESO_TONELADA', 0))
        r[uf]['fin'] += pf(row.get('VL_FINANCEIRO', 0))
    return dict(r)

def agg_by_permis(rows):
    r = defaultdict(lambda: {'ton': 0.0, 'fin': 0.0})
    for row in rows:
        p = (row.get('MAT_PERMIS', '') or row.get('NM_PERMIS', '')).strip()
        if p:
            r[p]['ton'] += pf(row.get('PESO_TONELADA', 0))
            r[p]['fin'] += pf(row.get('VL_FINANCEIRO', 0))
    return dict(r)

# ── OPENPYXL HELPERS ──────────────────────────────────────────────────────────
def find_row_by_label(ws, col, label, max_row=50):
    """Find row number where col has value == label (case insensitive)"""
    label_u = str(label).upper().strip()
    for row in ws.iter_rows(min_row=1, max_row=max_row):
        cell = row[col - 1]
        if cell.value and str(cell.value).upper().strip() == label_u:
            return cell.row
    return None

def find_rows_by_label(ws, col, label, max_row=50):
    """Find ALL rows where col has value == label"""
    label_u = str(label).upper().strip()
    result = []
    for row in ws.iter_rows(min_row=1, max_row=max_row):
        cell = row[col - 1]
        if cell.value and str(cell.value).upper().strip() == label_u:
            result.append(cell.row)
    return result

def set_val(ws, row, col, val):
    """Set numeric value, preserving format"""
    cell = ws.cell(row=row, column=col)
    if val is None:
        cell.value = None
    else:
        cell.value = float(val) if val != 0 else 0.0

def get_val(ws, row, col):
    v = ws.cell(row=row, column=col).value
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0

def var(a, b):
    return (a - b) / b if b else 0.0

def rston(ton, fin):
    return fin / ton if ton else 0.0

# ── SHEET UPDATERS ────────────────────────────────────────────────────────────

# MENSAL (cols B-O: FRUTAS ton/fin, LEGUMES ton/fin, VERDURAS ton/fin,
#         DIVERSOS ton/fin, PESCADOS ton/fin, FLORES ton/fin, ORGÂNICOS ton/fin)
# Col: A=1, B=2..O=15
MENSAL_GRUPOS = ['FRUTAS', 'LEGUMES', 'VERDURAS', 'DIVERSOS', 'PESCADOS', 'FLORES', 'ORGÂNICOS']
MENSAL_COLS = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]  # B..O

def update_mensal(ws, mes, etsp_data):
    mes_nome = MONTHS_PT[mes - 1]
    mes_row = find_row_by_label(ws, 1, mes_nome)
    total_row = find_row_by_label(ws, 1, 'TOTAL')
    if not mes_row:
        return

    # Write current month
    for i, g in enumerate(MENSAL_GRUPOS):
        tc = MENSAL_COLS[i * 2]      # ton col
        fc = MENSAL_COLS[i * 2 + 1]  # fin col
        set_val(ws, mes_row, tc, etsp_data.get(g, {}).get('ton', 0))
        set_val(ws, mes_row, fc, etsp_data.get(g, {}).get('fin', 0))

    # Recalculate TOTAL by summing all filled months
    if total_row:
        for i, g in enumerate(MENSAL_GRUPOS):
            tc = MENSAL_COLS[i * 2]
            fc = MENSAL_COLS[i * 2 + 1]
            st = sf = 0.0
            for r in range(6, total_row):  # rows 6..total_row-1
                lbl = ws.cell(row=r, column=1).value
                if lbl and str(lbl).upper().strip() in MONTHS_PT:
                    st += get_val(ws, r, tc)
                    sf += get_val(ws, r, fc)
            set_val(ws, total_row, tc, st)
            set_val(ws, total_row, fc, sf)


def update_comp_sheet(ws, mes, cur_ton, cur_fin, prev_year_monthly=None):
    """Update COMPARATIVO / INTERIOR / COMP_ETSP+INT style sheet.
    Two tables: R6-R18 monthly, R24-R36 YTD cumulative.
    """
    mes_rows = find_rows_by_label(ws, 1, MONTHS_PT[mes - 1], max_row=50)
    total_rows = find_rows_by_label(ws, 1, 'TOTAL', max_row=50)

    rston_cur = rston(cur_ton, cur_fin)

    # YTD prev year: sum months 1..mes from prev_year_monthly
    ytd_prev_ton = ytd_prev_fin = 0.0
    if prev_year_monthly:
        for m in range(1, mes + 1):
            if m in prev_year_monthly:
                ytd_prev_ton += prev_year_monthly[m]['ton']
                ytd_prev_fin += prev_year_monthly[m]['fin']

    # YTD current: cur_ton + sum of already-written months 1..mes-1 from table 1 (first occurrence)
    ytd_cur_ton = cur_ton
    ytd_cur_fin = cur_fin
    if mes_rows:
        first_mes_row = min(mes_rows)
        for r in range(6, first_mes_row):
            lbl = ws.cell(row=r, column=1).value
            if lbl and str(lbl).upper().strip() in MONTHS_PT:
                v_ton = get_val(ws, r, 2)
                v_fin = get_val(ws, r, 3)
                if v_ton: ytd_cur_ton += v_ton
                if v_fin: ytd_cur_fin += v_fin

    for i, mr in enumerate(mes_rows):
        is_ytd = (i > 0)
        if is_ytd:
            # Cumulative table
            set_val(ws, mr, 2, ytd_cur_ton)
            set_val(ws, mr, 3, ytd_cur_fin)
            set_val(ws, mr, 4, rston(ytd_cur_ton, ytd_cur_fin))
            if ytd_prev_ton:
                set_val(ws, mr, 5, ytd_prev_ton)
                set_val(ws, mr, 6, ytd_prev_fin)
                set_val(ws, mr, 7, rston(ytd_prev_ton, ytd_prev_fin))
                set_val(ws, mr, 8, var(ytd_cur_ton, ytd_prev_ton))
                set_val(ws, mr, 9, var(ytd_cur_fin, ytd_prev_fin))
                set_val(ws, mr, 10, var(rston(ytd_cur_ton, ytd_cur_fin), rston(ytd_prev_ton, ytd_prev_fin)))
        else:
            # Monthly table
            set_val(ws, mr, 2, cur_ton)
            set_val(ws, mr, 3, cur_fin)
            set_val(ws, mr, 4, rston_cur)
            # Prev year monthly for this month
            if prev_year_monthly and mes in prev_year_monthly:
                pt = prev_year_monthly[mes]['ton']
                pf_ = prev_year_monthly[mes]['fin']
                set_val(ws, mr, 5, pt)
                set_val(ws, mr, 6, pf_)
                set_val(ws, mr, 7, rston(pt, pf_))
                if pt:
                    set_val(ws, mr, 8, var(cur_ton, pt))
                    set_val(ws, mr, 9, var(cur_fin, pf_))
                    set_val(ws, mr, 10, var(rston_cur, rston(pt, pf_)))
            else:
                # Keep existing prev-year values, recalculate variation
                pt = get_val(ws, mr, 5)
                pf_ = get_val(ws, mr, 6)
                if pt:
                    set_val(ws, mr, 8, var(cur_ton, pt))
                    set_val(ws, mr, 9, var(cur_fin, pf_))
                    set_val(ws, mr, 10, var(rston_cur, rston(pt, pf_)))

    # Update TOTAL rows
    for i, total_row in enumerate(total_rows):
        start = total_rows[i - 1] if i > 0 else 0
        sb = sc = se = sf2 = 0.0
        for r in range(max(1, start + 1), total_row):
            lbl = ws.cell(row=r, column=1).value
            if not lbl or str(lbl).upper().strip() not in MONTHS_PT:
                continue
            sb += get_val(ws, r, 2)
            sc += get_val(ws, r, 3)
            se += get_val(ws, r, 5)
            sf2 += get_val(ws, r, 6)
        set_val(ws, total_row, 2, sb)
        set_val(ws, total_row, 3, sc)
        set_val(ws, total_row, 4, rston(sb, sc))
        if se:
            set_val(ws, total_row, 5, se)
            set_val(ws, total_row, 6, sf2)
            set_val(ws, total_row, 7, rston(se, sf2))
            set_val(ws, total_row, 8, var(sb, se))
            set_val(ws, total_row, 9, var(sc, sf2))
            set_val(ws, total_row, 10, var(rston(sb, sc), rston(se, sf2)))


def update_category_sheet(ws, mes, cur_ton, cur_fin, prev_year_monthly=None):
    """Category sheet (Frutas, Legumes, etc.) - same layout as comp"""
    update_comp_sheet(ws, mes, cur_ton, cur_fin, prev_year_monthly)


# GERAL / GERAL_AC - both have same col layout but GERAL has ORGÂNICOS, GERAL_AC doesn't
# Col layout for GERAL_JAN: B(2)=FRUTAS ton, C(3)=FRUTAS fin, D(4)=FRUTAS R$/ton
#   E(5)=LEGUMES ton, F(6)=fin, G(7)=R$/ton
#   H(8)=VERDURAS, I(9), J(10)
#   K(11)=DIVERSOS, L(12), M(13)
#   N(14)=FLORES, O(15), P(16)
#   Q(17)=PESCADOS, R(18), S(19)
#   T(20)=ORGÂNICOS, U(21), V(22)  ← only in GERAL_JAN
#   W(23)=TOTAL, X(24), Y(25)      ← GERAL_JAN
# For GERAL_AC (no ORGÂNICOS):
#   T(20)=TOTAL ton, U(21)=TOTAL fin, V(22)=TOTAL R$/ton

GERAL_GRUPOS = ['FRUTAS', 'LEGUMES', 'VERDURAS', 'DIVERSOS', 'FLORES', 'PESCADOS', 'ORGÂNICOS']
GERAL_COL_START = 2  # B = col 2

def _geral_col(g_idx, offset=0):
    """Column for grupo g_idx (0-based), offset 0=ton, 1=fin, 2=R$/ton"""
    return GERAL_COL_START + g_idx * 3 + offset

def update_geral(ws, etsp_data, interior_unit_data, has_organicos=True):
    """Update GERAL_XXX monthly snapshot sheet"""
    # Find rows by unit name
    unit_row = {}
    total_int_row = etsp_row = total_all_row = None
    for row in ws.iter_rows(min_row=5, max_row=30):
        lbl = row[0].value
        if not lbl:
            continue
        ls = str(lbl).upper().strip()
        for code, uname in SHEET_TO_UNIT.items():
            if ls == uname:
                unit_row[code] = row[0].row
        if ls == 'TOTAL INTERIOR':
            total_int_row = row[0].row
        elif ls == 'ETSP':
            etsp_row = row[0].row
        elif ls == 'TOTAL ENTREPOSTOS':
            total_all_row = row[0].row

    grupos = GERAL_GRUPOS if has_organicos else GRUPOS_INT
    n_grupos = len(grupos)
    total_col = GERAL_COL_START + n_grupos * 3  # W for GERAL_JAN, T for GERAL_AC

    def write_row(r, data):
        for i, g in enumerate(grupos):
            t = data.get(g, {}).get('ton', 0)
            f = data.get(g, {}).get('fin', 0)
            c = GERAL_COL_START + i * 3
            set_val(ws, r, c, t)
            set_val(ws, r, c + 1, f)
            set_val(ws, r, c + 2, rston(t, f))
        tT = sum(data.get(g, {}).get('ton', 0) for g in grupos)
        fT = sum(data.get(g, {}).get('fin', 0) for g in grupos)
        set_val(ws, r, total_col, tT)
        set_val(ws, r, total_col + 1, fT)
        set_val(ws, r, total_col + 2, rston(tT, fT))

    # Write interior units
    for code in SHEET_TO_UNIT:
        r = unit_row.get(code)
        if r and code in interior_unit_data:
            write_row(r, interior_unit_data[code])

    # Sum TOTAL INTERIOR from updated values
    if total_int_row:
        sums = {g: {'ton': 0.0, 'fin': 0.0} for g in grupos}
        for code in SHEET_TO_UNIT:
            r = unit_row.get(code)
            if not r:
                continue
            for i, g in enumerate(grupos):
                c = GERAL_COL_START + i * 3
                sums[g]['ton'] += get_val(ws, r, c)
                sums[g]['fin'] += get_val(ws, r, c + 1)
        write_row(total_int_row, sums)

    # Write ETSP
    if etsp_row:
        write_row(etsp_row, etsp_data)

    # TOTAL ENTREPOSTOS = ETSP + TOTAL INTERIOR
    if total_all_row and etsp_row and total_int_row:
        sums2 = {g: {'ton': 0.0, 'fin': 0.0} for g in grupos}
        for src_r in [etsp_row, total_int_row]:
            for i, g in enumerate(grupos):
                c = GERAL_COL_START + i * 3
                sums2[g]['ton'] += get_val(ws, src_r, c)
                sums2[g]['fin'] += get_val(ws, src_r, c + 1)
        write_row(total_all_row, sums2)


def update_geral_ac(ws, etsp_data, interior_unit_data):
    """GERAL_AC: YTD accumulated. Reads existing template values and adds current month.
    GERAL_AC has NO ORGÂNICOS column."""
    unit_row = {}
    total_int_row = etsp_row = total_all_row = None
    for row in ws.iter_rows(min_row=5, max_row=30):
        lbl = row[0].value
        if not lbl:
            continue
        ls = str(lbl).upper().strip()
        for code, uname in SHEET_TO_UNIT.items():
            if ls == uname:
                unit_row[code] = row[0].row
        if ls == 'TOTAL INTERIOR':
            total_int_row = row[0].row
        elif ls == 'ETSP':
            etsp_row = row[0].row
        elif ls == 'TOTAL ENTREPOSTOS':
            total_all_row = row[0].row

    grupos = GRUPOS_INT  # no ORGÂNICOS in GERAL_AC
    n_grupos = len(grupos)
    total_col = GERAL_COL_START + n_grupos * 3

    def read_row(r):
        d = {}
        for i, g in enumerate(grupos):
            c = GERAL_COL_START + i * 3
            d[g] = {'ton': get_val(ws, r, c), 'fin': get_val(ws, r, c + 1)}
        return d

    def write_row_ac(r, data):
        for i, g in enumerate(grupos):
            t = data.get(g, {}).get('ton', 0)
            f = data.get(g, {}).get('fin', 0)
            c = GERAL_COL_START + i * 3
            set_val(ws, r, c, t)
            set_val(ws, r, c + 1, f)
            set_val(ws, r, c + 2, rston(t, f))
        tT = sum(data.get(g, {}).get('ton', 0) for g in grupos)
        fT = sum(data.get(g, {}).get('fin', 0) for g in grupos)
        set_val(ws, r, total_col, tT)
        set_val(ws, r, total_col + 1, fT)
        set_val(ws, r, total_col + 2, rston(tT, fT))

    # Interior units: template YTD + current month
    for code in SHEET_TO_UNIT:
        r = unit_row.get(code)
        if not r:
            continue
        ytd = read_row(r)
        new_d = interior_unit_data.get(code, {})
        merged = {g: {'ton': ytd[g]['ton'] + new_d.get(g, {}).get('ton', 0),
                      'fin': ytd[g]['fin'] + new_d.get(g, {}).get('fin', 0)}
                  for g in grupos}
        write_row_ac(r, merged)

    # ETSP: template YTD + current month
    if etsp_row:
        ytd_etsp = read_row(etsp_row)
        merged_etsp = {g: {'ton': ytd_etsp[g]['ton'] + etsp_data.get(g, {}).get('ton', 0),
                            'fin': ytd_etsp[g]['fin'] + etsp_data.get(g, {}).get('fin', 0)}
                       for g in grupos}
        write_row_ac(etsp_row, merged_etsp)

    # TOTAL INTERIOR: sum all interior units from updated sheet
    if total_int_row:
        sums = {g: {'ton': 0.0, 'fin': 0.0} for g in grupos}
        for code in SHEET_TO_UNIT:
            r = unit_row.get(code)
            if not r:
                continue
            for i, g in enumerate(grupos):
                c = GERAL_COL_START + i * 3
                sums[g]['ton'] += get_val(ws, r, c)
                sums[g]['fin'] += get_val(ws, r, c + 1)
        write_row_ac(total_int_row, sums)

    # TOTAL ENTREPOSTOS
    if total_all_row and etsp_row and total_int_row:
        sums2 = {g: {'ton': 0.0, 'fin': 0.0} for g in grupos}
        for src_r in [etsp_row, total_int_row]:
            for i, g in enumerate(grupos):
                c = GERAL_COL_START + i * 3
                sums2[g]['ton'] += get_val(ws, src_r, c)
                sums2[g]['fin'] += get_val(ws, src_r, c + 1)
        write_row_ac(total_all_row, sums2)


def update_interior_unit(ws, mes, data):
    """Interior unit sheet (Ceara, Cearb, etc.)
    Cols: B(2)=FRUTAS ton, C(3)=fin, D(4)=LEGUMES ton, E(5)=fin,
          F(6)=VERDURAS ton, G(7)=fin, H(8)=DIVERSOS ton, I(9)=fin,
          J(10)=FLORES ton, K(11)=fin, L(12)=PESCADOS ton, M(13)=fin,
          N(14)=TOTAL ton, O(15)=fin
    """
    INT_GRUPOS = GRUPOS_INT
    INT_COLS = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]  # pairs: B/C, D/E, F/G, H/I, J/K, L/M

    mes_nome = MONTHS_PT[mes - 1]
    mes_row = find_row_by_label(ws, 1, mes_nome)
    total_row = find_row_by_label(ws, 1, 'TOTAL')
    if not mes_row:
        return

    for i, g in enumerate(INT_GRUPOS):
        tc = INT_COLS[i * 2]
        fc = INT_COLS[i * 2 + 1]
        set_val(ws, mes_row, tc, data.get(g, {}).get('ton', 0))
        set_val(ws, mes_row, fc, data.get(g, {}).get('fin', 0))

    # TOTAL cols N=14, O=15
    tt = sum(data.get(g, {}).get('ton', 0) for g in INT_GRUPOS)
    tf = sum(data.get(g, {}).get('fin', 0) for g in INT_GRUPOS)
    set_val(ws, mes_row, 14, tt)
    set_val(ws, mes_row, 15, tf)

    if total_row:
        for i, g in enumerate(INT_GRUPOS):
            tc = INT_COLS[i * 2]
            fc = INT_COLS[i * 2 + 1]
            st = sf = 0.0
            for r in range(6, total_row):
                lbl = ws.cell(row=r, column=1).value
                if lbl and str(lbl).upper().strip() in MONTHS_PT:
                    st += get_val(ws, r, tc)
                    sf += get_val(ws, r, fc)
            set_val(ws, total_row, tc, st)
            set_val(ws, total_row, fc, sf)
        # TOTAL cols
        stt = sum(get_val(ws, total_row, INT_COLS[i * 2]) for i in range(len(INT_GRUPOS)))
        stf = sum(get_val(ws, total_row, INT_COLS[i * 2 + 1]) for i in range(len(INT_GRUPOS)))
        set_val(ws, total_row, 14, stt)
        set_val(ws, total_row, 15, stf)


def update_setores(ws, etsp_data):
    """SETORES: accumulate YTD by reading existing C values and adding current month.
    Table 1 (tons) rows ~4-10, TOTAL row.
    Table 2 (fin) rows ~16-22, TOTAL row.
    A col has ordinal ('1º'), B col has PRODUTO name string.
    """
    total_rows = []
    for row in ws.iter_rows(min_row=1, max_row=30):
        lbl = row[0].value
        if lbl and str(lbl).upper().strip() == 'TOTAL':
            total_rows.append(row[0].row)

    if not total_rows:
        return

    first_total = total_rows[0]
    second_total = total_rows[1] if len(total_rows) > 1 else 9999

    # YTD: read existing C values from ton table + add current month
    ytd_ton = defaultdict(float)
    ytd_fin = defaultdict(float)

    for row in ws.iter_rows(min_row=1, max_row=first_total - 1):
        b = row[1].value  # produto name
        c = row[2].value  # current YTD value
        if b and isinstance(b, str) and str(row[0].value or '').strip().endswith('º'):
            g = b.strip().upper()
            ytd_ton[g] += float(c or 0)

    for row in ws.iter_rows(min_row=first_total + 1, max_row=second_total - 1):
        b = row[1].value
        c = row[2].value
        if b and isinstance(b, str) and str(row[0].value or '').strip().endswith('º'):
            g = b.strip().upper()
            ytd_fin[g] += float(c or 0)

    # Add current month
    for g, d in etsp_data.items():
        gu = g.upper()
        ytd_ton[gu] += d.get('ton', 0)
        ytd_fin[gu] += d.get('fin', 0)

    total_ton = sum(ytd_ton.values())
    total_fin = sum(ytd_fin.values())

    # Write ton table
    for row in ws.iter_rows(min_row=1, max_row=first_total - 1):
        b = row[1].value
        if b and isinstance(b, str) and str(row[0].value or '').strip().endswith('º'):
            g = b.strip().upper()
            row[2].value = ytd_ton[g]
            row[3].value = ytd_ton[g] / total_ton if total_ton else 0

    if total_rows:
        ws.cell(total_rows[0], 3).value = total_ton

    # Write fin table
    for row in ws.iter_rows(min_row=first_total + 1, max_row=second_total - 1):
        b = row[1].value
        if b and isinstance(b, str) and str(row[0].value or '').strip().endswith('º'):
            g = b.strip().upper()
            row[2].value = ytd_fin[g]
            row[3].value = ytd_fin[g] / total_fin if total_fin else 0

    if len(total_rows) > 1:
        ws.cell(total_rows[1], 3).value = total_fin


def update_rank_sheet(ws, new_prod_data):
    """Product ranking sheet. Accumulates YTD from existing data.
    Two tables: TON (before first TOTAL), FIN (between first and second TOTAL).
    A=ordinal, B=produto, C=value, D=partic%.
    """
    total_rnums = []
    for row in ws.iter_rows(min_row=1, max_row=200):
        a = row[0].value
        if a and str(a).upper().strip() == 'TOTAL':
            total_rnums.append(row[0].row)

    first_t = total_rnums[0] if total_rnums else 9999
    second_t = total_rnums[1] if len(total_rnums) > 1 else 9999

    # Read existing YTD from TON table
    ytd = defaultdict(lambda: {'ton': 0.0, 'fin': 0.0})
    for row in ws.iter_rows(min_row=1, max_row=first_t - 1):
        a = row[0].value
        b = row[1].value
        c = row[2].value
        if a and str(a).strip().endswith('º') and b and isinstance(b, str):
            ytd[b.strip().upper()]['ton'] += float(c or 0)

    # Add current month
    for prod, d in new_prod_data.items():
        ytd[prod]['ton'] += d.get('ton', 0)
        ytd[prod]['fin'] += d.get('fin', 0)

    total_ton = sum(v['ton'] for v in ytd.values())
    total_fin = sum(v['fin'] for v in ytd.values())
    sorted_ton = sorted(ytd.items(), key=lambda x: -x[1]['ton'])
    sorted_fin = sorted(ytd.items(), key=lambda x: -x[1]['fin'])

    # Collect data rows
    ton_rows = [row for row in ws.iter_rows(min_row=1, max_row=first_t - 1)
                if row[0].value and str(row[0].value).strip().endswith('º')]
    fin_rows = [row for row in ws.iter_rows(min_row=first_t + 1, max_row=second_t - 1)
                if row[0].value and str(row[0].value).strip().endswith('º')]

    # Write TON table
    acum = 0.0
    for i, row in enumerate(ton_rows):
        if i < len(sorted_ton):
            prod, vals = sorted_ton[i]
            t = vals['ton']
            p = t / total_ton if total_ton else 0
            acum += p
            row[1].value = prod
            row[2].value = t
            row[3].value = p
            if len(row) > 4 and row[4].value is not None:
                row[4].value = acum
        else:
            row[2].value = 0
            row[3].value = 0

    if total_rnums:
        ws.cell(total_rnums[0], 3).value = total_ton

    # Write FIN table
    acum = 0.0
    for i, row in enumerate(fin_rows):
        if i < len(sorted_fin):
            prod, vals = sorted_fin[i]
            f = vals['fin']
            p = f / total_fin if total_fin else 0
            acum += p
            row[1].value = prod
            row[2].value = f
            row[3].value = p
            if len(row) > 4 and row[4].value is not None:
                row[4].value = acum
        else:
            row[2].value = 0
            row[3].value = 0

    if len(total_rnums) > 1:
        ws.cell(total_rnums[1], 3).value = total_fin


def update_origem_rank(ws, new_data):
    """País/UF ranking sheet. Same structure as rank_sheet."""
    update_rank_sheet(ws, new_data)


def update_perm_sheet(ws, etsp_rows_csv, grupo_filter=None):
    """Permissionário sheet.
    Structure: A=threshold(0.25/0.5/0.75/1.0), B=n_permis, C=ton_or_fin
    Two tables: ton then fin.
    Read existing YTD from threshold data, add current month, recalculate.
    """
    # Filter rows
    if grupo_filter:
        filtered = [r for r in etsp_rows_csv
                    if GRUPO_MAP.get((r.get('GRUPO_PRODUTO', '') or '').strip().upper()) == grupo_filter]
    else:
        filtered = etsp_rows_csv

    # Current month permis data
    mes_permis = agg_by_permis(filtered)

    # Read existing YTD from ton table (rows with A = 0.25/0.5/0.75/1.0)
    thresholds = [0.25, 0.5, 0.75, 1.0]
    ton_rows_info = []  # (threshold, row_num)
    fin_rows_info = []
    switched = False
    for row in ws.iter_rows(min_row=1, max_row=25):
        a = row[0].value
        try:
            v = float(a)
            if v in thresholds:
                if not switched:
                    ton_rows_info.append((v, row[0].row))
                else:
                    fin_rows_info.append((v, row[0].row))
        except:
            pass
        # Detect switch to fin table
        for cell in row:
            if cell.value and isinstance(cell.value, str) and 'FINANC' in cell.value.upper():
                switched = True

    # Read existing YTD from ton table's C column
    prev_ton_ytd = {}
    prev_fin_ytd = {}
    for thresh, r in ton_rows_info:
        prev_ton_ytd[thresh] = (int(get_val(ws, r, 2) or 0), get_val(ws, r, 3))
    for thresh, r in fin_rows_info:
        prev_fin_ytd[thresh] = (int(get_val(ws, r, 2) or 0), get_val(ws, r, 3))

    # Reconstruct virtual permis from previous months' data
    all_permis = defaultdict(lambda: {'ton': 0.0, 'fin': 0.0})

    # Add virtual permis from previous thresholds
    if prev_ton_ytd.get(1.0):
        total_n_prev, total_ton_prev = prev_ton_ytd[1.0]
        if total_n_prev > 0 and total_ton_prev > 0:
            prev_n = 0
            prev_cum_ton = 0.0
            prev_cum_fin = 0.0
            for t in thresholds:
                if t not in prev_ton_ytd:
                    continue
                n, cum_ton = prev_ton_ytd[t]
                _, cum_fin = prev_fin_ytd.get(t, (n, 0.0))
                bucket_n = n - prev_n
                bucket_ton = cum_ton - prev_cum_ton
                bucket_fin = cum_fin - prev_cum_fin
                if bucket_n > 0:
                    avg_ton = bucket_ton / bucket_n
                    avg_fin = bucket_fin / bucket_n
                    for i in range(bucket_n):
                        pid = f'_prev_{t}_{i}'
                        all_permis[pid]['ton'] += avg_ton
                        all_permis[pid]['fin'] += avg_fin
                prev_n = n
                prev_cum_ton = cum_ton
                prev_cum_fin = cum_fin

    # Add current month permis
    for pid, vals in mes_permis.items():
        all_permis[pid]['ton'] += vals.get('ton', 0)
        all_permis[pid]['fin'] += vals.get('fin', 0)

    sorted_ton = sorted(all_permis.items(), key=lambda x: -x[1]['ton'])
    sorted_fin = sorted(all_permis.items(), key=lambda x: -x[1]['fin'])
    total_ton = sum(v['ton'] for v in all_permis.values())
    total_fin = sum(v['fin'] for v in all_permis.values())

    def calc_thresh(sorted_items, key, total, t):
        running = 0.0
        for i, (_, v) in enumerate(sorted_items):
            running += v[key]
            if running >= total * t:
                return i + 1, running
        return len(sorted_items), total

    for thresh, r in ton_rows_info:
        n, cum = calc_thresh(sorted_ton, 'ton', total_ton, thresh)
        set_val(ws, r, 2, n)
        set_val(ws, r, 3, cum)

    for thresh, r in fin_rows_info:
        n, cum = calc_thresh(sorted_fin, 'fin', total_fin, thresh)
        set_val(ws, r, 2, n)
        set_val(ws, r, 3, cum)


def update_ranking_units(ws, units_ton, units_fin):
    """RANKING_M_ETSP / RANKING_M_INT.
    Two sub-tables: ton (rows ~5..TOTAL) and fin (rows after TOTAL).
    C=value, D=partic%, E=acumul%.
    """
    total_rows = []
    for row in ws.iter_rows(min_row=1, max_row=40):
        a = row[0].value
        if a and 'TOTAL' in str(a).upper():
            total_rows.append(row[0].row)

    def write_ranking(rows_iter, units, total_val, start_after, end_before):
        data_rows = [row for row in rows_iter
                     if start_after < row[0].row < end_before
                     and row[0].value
                     and str(row[0].value or '').strip().endswith('º')]
        acum = 0.0
        for i, row in enumerate(data_rows):
            if i < len(units):
                name, val = units[i]
                p = val / total_val if total_val else 0
                acum += p
                row[1].value = name
                row[2].value = val
                row[3].value = p
                if len(row) > 4:
                    row[4].value = acum

    all_rows = list(ws.iter_rows(min_row=1, max_row=40))
    mid = total_rows[0] if total_rows else 9999
    end = total_rows[1] if len(total_rows) > 1 else 9999

    total_ton_val = sum(v for _, v in units_ton)
    total_fin_val = sum(v for _, v in units_fin)

    write_ranking(all_rows, units_ton, total_ton_val, 0, mid)
    write_ranking(all_rows, units_fin, total_fin_val, mid, end)

    # Write total rows
    if total_rows:
        ws.cell(total_rows[0], 3).value = total_ton_val
    if len(total_rows) > 1:
        ws.cell(total_rows[1], 3).value = total_fin_val


def update_ranking_units_ac(ws, units_ton, units_fin):
    """RANKING_M_AC_ETSP / RANKING_MÊS_AC - same structure but YTD."""
    update_ranking_units(ws, units_ton, units_fin)


def update_titles_in_sheet(ws, mes, ano):
    """Update year/month references in text cells"""
    mes_nome = MONTHS_PT[mes - 1]
    mes_short = MONTHS_SHORT[mes - 1]
    prev_nome = MONTHS_PT[mes - 2] if mes > 1 else MONTHS_PT[11]
    prev_short = MONTHS_SHORT[mes - 2] if mes > 1 else MONTHS_SHORT[11]
    ano_str = str(ano)
    prev_ano_str = str(ano - 1)

    patterns = [
        (f'{prev_nome} - {ano_str}', f'{mes_nome} - {ano_str}'),
        (f'({prev_nome})', f'({mes_nome})'),
        (f'{prev_nome}/{ano_str}', f'{mes_nome}/{ano_str}'),
        (f'{prev_nome}/{ano_str[-2:]}', f'{mes_nome}/{ano_str[-2:]}'),
        (f'( {prev_short} )', f'( {mes_short} )'),
        (f'({prev_short})', f'({mes_short})'),
        (f'{prev_nome} {ano_str}', f'{mes_nome} {ano_str}'),
        (f'JAN - {prev_short} {ano_str}', f'JAN - {mes_short} {ano_str}'),
        (f'JAN - {prev_short}', f'JAN - {mes_short}'),
        (f'JAN-{prev_short}', f'JAN-{mes_short}'),
    ]

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                v = cell.value
                for old, new in patterns:
                    v = v.replace(old, new)
                if v != cell.value:
                    cell.value = v


# ── MAIN ENTRY POINT ──────────────────────────────────────────────────────────

def process_bytes(template_bytes, csv_data_list, mes, ano, prev_year_bytes=None):
    """
    template_bytes   : bytes of previous month's XLSX (current year)
    csv_data_list    : list of {'filename': str, 'content': str (latin1 text)}
    mes, ano         : int
    prev_year_bytes  : optional bytes of previous year's December XLSX
    Returns: bytes of the new XLSX
    """
    print(f"Processando Balanço {MONTHS_PT[mes-1]} {ano}")

    # 1. Parse CSVs
    etsp_rows = []
    interior_rows_by_sheet = defaultdict(list)

    for item in csv_data_list:
        rows = parse_csv_content(item['content'])
        if not rows:
            continue
        code = detect_unit(rows)
        if not code:
            print(f"  ⚠ CSV '{item['filename']}': unidade não reconhecida")
            continue
        if code == 'ETSP':
            etsp_rows = rows
        else:
            interior_rows_by_sheet[code] = rows
        print(f"  CSV '{item['filename']}': {len(rows)} linhas → {code}")

    etsp_data = agg_by_grupo(etsp_rows)
    interior_unit_data = {c: agg_by_grupo(r) for c, r in interior_rows_by_sheet.items()}

    etsp_ton = sum(etsp_data[g]['ton'] for g in GRUPOS)
    etsp_fin = sum(etsp_data[g]['fin'] for g in GRUPOS)
    int_ton = sum(sum(interior_unit_data.get(c, {}).get(g, {}).get('ton', 0) for g in GRUPOS)
                  for c in SHEET_TO_UNIT)
    int_fin = sum(sum(interior_unit_data.get(c, {}).get(g, {}).get('fin', 0) for g in GRUPOS)
                  for c in SHEET_TO_UNIT)
    all_int_rows = [r for rows in interior_rows_by_sheet.values() for r in rows]

    print(f"  ETSP: {etsp_ton:,.0f} ton, R$ {etsp_fin:,.0f}")

    # 2. Extract prev year data
    prev_year_data = {}
    if prev_year_bytes:
        print("  Lendo dados do ano anterior...")
        try:
            prev_wb = load_workbook(io.BytesIO(prev_year_bytes), read_only=True, data_only=True)
            comp_sheets = ['COMPARATIVO', 'FRUTAS', 'Frutas', 'FRUTAS ', 'Frutas ',
                           'LEGUMES', 'Legumes', 'LEGUMES ', 'Legumes ',
                           'VERDURAS', 'Verduras', 'VERDURAS ', 'Verduras ',
                           'DIVERSOS', 'Diversos', 'DIVERS', 'FLORES', 'Flores', 'FLORES ',
                           'PESCADO', 'PESCADOS', 'Pescado', 'Pescados',
                           'ORGÂNICOS', 'Orgânicos', 'ORGANICOS', 'Orgânico',
                           'INTERIOR', 'COMP_ETSP+INT',
                           'Ceara1', 'Cearb1', 'Cebau1', 'Cefra1', 'Cemar1', 'Cepir1',
                           'Cepre1', 'Cerib1', 'Cesjc1', 'Cesjr1', 'Cesor1', 'Cegua1']
            for sn in prev_wb.sheetnames:
                ws_p = prev_wb[sn]
                monthly = {}
                for row in ws_p.iter_rows(max_row=50, values_only=True):
                    lbl = row[0] if row else None
                    if lbl and str(lbl).upper().strip() in MONTHS_PT:
                        mi = MONTHS_PT.index(str(lbl).upper().strip()) + 1
                        b = row[1] if len(row) > 1 else None
                        c = row[2] if len(row) > 2 else None
                        try:
                            t = float(b) if b else 0
                            f = float(c) if c else 0
                            if t or f:
                                monthly[mi] = {'ton': t, 'fin': f}
                        except:
                            pass
                if monthly:
                    prev_year_data[sn] = monthly
                    # Also store normalized keys
                    sn_norm = sn.strip().upper()
                    prev_year_data[sn_norm] = monthly
            prev_wb.close()
            print(f"  Ano anterior: {len(set(prev_year_data.keys()))} sheets com dados mensais")
        except Exception as e:
            print(f"  ⚠ Erro ao ler ano anterior: {e}")

    def py(key):
        """Get prev_year_monthly for a sheet key"""
        for k in [key, key.strip(), key.upper().strip()]:
            if k in prev_year_data:
                return prev_year_data[k]
        return None

    # 3. Load template workbook (with full formatting)
    wb = load_workbook(io.BytesIO(template_bytes))
    sheets = {sn.strip(): sn for sn in wb.sheetnames}  # normalized → real name

    def gs(name):
        """Get sheet by name (try various capitalizations)"""
        # Direct match
        if name in wb.sheetnames:
            return wb[name]
        # Strip match
        real = sheets.get(name.strip())
        if real:
            return wb[real]
        # Case-insensitive
        nu = name.strip().upper()
        for sn in wb.sheetnames:
            if sn.strip().upper() == nu:
                return wb[sn]
        return None

    # 4. MENSAL
    ws = gs('MENSAL')
    if ws:
        update_mensal(ws, mes, etsp_data)
        print("  ✓ MENSAL")

    # 5. COMPARATIVO (ETSP monthly + YTD)
    ws = gs('COMPARATIVO')
    if ws:
        update_comp_sheet(ws, mes, etsp_ton, etsp_fin, py('COMPARATIVO'))
        print("  ✓ COMPARATIVO")

    # 6. GERAL_MMM (monthly snapshot)
    prev_short = MONTHS_SHORT[mes - 2].upper() if mes > 1 else MONTHS_SHORT[11].upper()
    cur_short = MONTHS_SHORT[mes - 1].upper()
    old_geral = f'GERAL_{prev_short}'
    new_geral = f'GERAL_{cur_short}'

    ws = gs(old_geral)
    if ws:
        update_geral(ws, etsp_data, interior_unit_data, has_organicos=True)
        # Rename the sheet
        ws.title = new_geral
        print(f"  ✓ {old_geral} → {new_geral}")
    else:
        ws = gs(new_geral)
        if ws:
            update_geral(ws, etsp_data, interior_unit_data, has_organicos=True)
            print(f"  ✓ {new_geral}")

    # 7. GERAL_AC (YTD accumulated)
    ws = gs('GERAL_AC')
    if ws:
        update_geral_ac(ws, etsp_data, interior_unit_data)
        print("  ✓ GERAL_AC")

    # 8. COMP_ETSP+INT
    ws = gs('COMP_ETSP+INT')
    if ws:
        update_comp_sheet(ws, mes, etsp_ton + int_ton, etsp_fin + int_fin, py('COMP_ETSP+INT'))
        print("  ✓ COMP_ETSP+INT")

    # 9. INTERIOR
    ws = gs('INTERIOR')
    if ws:
        update_comp_sheet(ws, mes, int_ton, int_fin, py('INTERIOR'))
        print("  ✓ INTERIOR")

    # 10. SETORES (ETSP sectors YTD)
    ws = gs('SETORES')
    if ws:
        update_setores(ws, etsp_data)
        print("  ✓ SETORES")

    # 11. Category sheets (Frutas, Legumes, Verduras, Diversos, Flores, Pescados/o, Orgânicos)
    # Sheet names vary between files - try multiple variants
    cat_map = [
        (['Frutas', 'FRUTAS', 'Frutas '], 'FRUTAS'),
        (['Legumes', 'LEGUMES', 'Legumes '], 'LEGUMES'),
        (['Verduras', 'VERDURAS', 'Verduras '], 'VERDURAS'),
        (['Diversos', 'DIVERSOS'], 'DIVERSOS'),
        (['Flores', 'FLORES', 'Flores '], 'FLORES'),
        (['Pescados', 'Pescado', 'PESCADOS', 'PESCADO'], 'PESCADOS'),
        (['Orgânicos', 'ORGÂNICOS', 'Orgânico', 'ORGANICOS', 'Orgânico '], 'ORGÂNICOS'),
    ]
    for names, g in cat_map:
        ws = None
        sn_found = None
        for n in names:
            ws = gs(n)
            if ws:
                sn_found = n
                break
        if ws:
            cur_t = etsp_data.get(g, {}).get('ton', 0)
            cur_f = etsp_data.get(g, {}).get('fin', 0)
            # Try multiple prev_year keys
            py_data = py(sn_found) or py(g) or py(names[-1])
            update_category_sheet(ws, mes, cur_t, cur_f, py_data)
    print("  ✓ Category sheets")

    # 12. Rank sheets per category
    rank_map = [
        (['Rank(F)', 'Rank(F) '], 'FRUTAS'),
        (['Rank(L)', 'Rank(L) '], 'LEGUMES'),
        (['Rank(V)', 'Rank(V) '], 'VERDURAS'),
        (['Rank(D)', 'Rank(D) '], 'DIVERSOS'),
        (['Rank(Fl)', 'Rank(FL)', 'Rank(FI)', 'Rank(Fl) '], 'FLORES'),
        (['Rank(P)', 'Rank(P) '], 'PESCADOS'),
        (['Rank(O)', 'Rank(O) '], 'ORGÂNICOS'),
    ]
    for names, g in rank_map:
        for n in names:
            ws = gs(n)
            if ws:
                update_rank_sheet(ws, agg_by_produto(etsp_rows, g))
                break
    print("  ✓ Rank sheets")

    # 13. Perm sheets
    perm_map = [
        (['Perm(F)', 'Perm(F) '], 'FRUTAS'),
        (['Perm(L)', 'Perm(L) '], 'LEGUMES'),
        (['Perm(V)', 'Perm(V) '], 'VERDURAS'),
        (['Perm(D)', 'Perm(D) '], 'DIVERSOS'),
        (['Perm(Fl)', 'Perm(FL)', 'Perm(Fl) '], 'FLORES'),
        (['Perm(P)', 'Perm(P) '], 'PESCADOS'),
        (['Perm(O)', 'PERM(O)', 'Perm(O) '], 'ORGÂNICOS'),
    ]
    for names, g in perm_map:
        for n in names:
            ws = gs(n)
            if ws:
                update_perm_sheet(ws, etsp_rows, g)
                break
    print("  ✓ Perm sheets")

    # 14. RankGeral PRODUTOS (all ETSP products)
    ws = gs('RankGeral PRODUTOS')
    if ws:
        update_rank_sheet(ws, agg_by_produto(etsp_rows))
        print("  ✓ RankGeral PRODUTOS")

    # 15. Países
    ws = gs('Países') or gs('Países (2)') or gs('Paises')
    if ws:
        update_origem_rank(ws, agg_by_pais(etsp_rows))
        print("  ✓ Países")

    # 16. UF (ETSP state ranking)
    ws = gs('UF')
    if ws:
        update_origem_rank(ws, agg_by_uf(etsp_rows))
        print("  ✓ UF")

    # 17. Interior unit sheets (Ceara, Cearb, ...)
    cnt = 0
    for code in SHEET_TO_UNIT:
        if code not in interior_unit_data:
            continue
        ws = gs(code)
        if ws:
            update_interior_unit(ws, mes, interior_unit_data[code])
            cnt += 1
        # CompX sheet (Ceara1, etc.)
        ws1 = gs(code + '1')
        if ws1:
            t = sum(interior_unit_data[code].get(g, {}).get('ton', 0) for g in GRUPOS_INT)
            f = sum(interior_unit_data[code].get(g, {}).get('fin', 0) for g in GRUPOS_INT)
            update_comp_sheet(ws1, mes, t, f, py(code + '1'))
    print(f"  ✓ {cnt} interior unit sheets")

    # 18. RANKING_M_ETSP (monthly ranking all units)
    units_ton = sorted(
        [(SHEET_TO_UNIT.get(c, c), sum(interior_unit_data.get(c, {}).get(g, {}).get('ton', 0) for g in GRUPOS))
         for c in SHEET_TO_UNIT]
        + [('ETSP', etsp_ton)], key=lambda x: -x[1])
    units_fin = sorted(
        [(SHEET_TO_UNIT.get(c, c), sum(interior_unit_data.get(c, {}).get(g, {}).get('fin', 0) for g in GRUPOS))
         for c in SHEET_TO_UNIT]
        + [('ETSP', etsp_fin)], key=lambda x: -x[1])

    ws = gs('RANKING_M_ETSP')
    if ws:
        update_ranking_units(ws, units_ton, units_fin)
        print("  ✓ RANKING_M_ETSP")

    # 19. RANKING_M_INT (interior units only)
    int_u_ton = sorted(
        [(SHEET_TO_UNIT.get(c, c), sum(interior_unit_data.get(c, {}).get(g, {}).get('ton', 0) for g in GRUPOS))
         for c in SHEET_TO_UNIT], key=lambda x: -x[1])
    int_u_fin = sorted(
        [(SHEET_TO_UNIT.get(c, c), sum(interior_unit_data.get(c, {}).get(g, {}).get('fin', 0) for g in GRUPOS))
         for c in SHEET_TO_UNIT], key=lambda x: -x[1])

    ws = gs('RANKING_M_INT')
    if ws:
        update_ranking_units(ws, int_u_ton, int_u_fin)
        print("  ✓ RANKING_M_INT")

    # 20. RANKING_M_AC_ETSP (YTD accumulated ranking all units)
    # Read existing YTD from sheet, add current month
    ws = gs('RANKING_M_AC_ETSP')
    if ws:
        # Read existing YTD from ton table
        total_rnums_ac = []
        for row in ws.iter_rows(min_row=1, max_row=40):
            a = row[0].value
            if a and 'TOTAL' in str(a).upper():
                total_rnums_ac.append(row[0].row)
        first_t_ac = total_rnums_ac[0] if total_rnums_ac else 9999
        second_t_ac = total_rnums_ac[1] if len(total_rnums_ac) > 1 else 9999

        ytd_ton_ac = defaultdict(float)
        ytd_fin_ac = defaultdict(float)
        for row in ws.iter_rows(min_row=1, max_row=first_t_ac - 1):
            a, b, c = row[0].value, row[1].value, row[2].value
            if a and str(a).strip().endswith('º') and b and isinstance(b, str):
                ytd_ton_ac[b.strip()] += float(c or 0)
        for row in ws.iter_rows(min_row=first_t_ac + 1, max_row=second_t_ac - 1):
            a, b, c = row[0].value, row[1].value, row[2].value
            if a and str(a).strip().endswith('º') and b and isinstance(b, str):
                ytd_fin_ac[b.strip()] += float(c or 0)

        # Add current month
        for name, val in units_ton:
            ytd_ton_ac[name] += val
        for name, val in units_fin:
            ytd_fin_ac[name] += val

        ac_units_ton = sorted(ytd_ton_ac.items(), key=lambda x: -x[1])
        ac_units_fin = sorted(ytd_fin_ac.items(), key=lambda x: -x[1])
        update_ranking_units(ws, ac_units_ton, ac_units_fin)
        print("  ✓ RANKING_M_AC_ETSP")

    # 21. RANKING_MÊS_AC (interior YTD)
    ws = gs('RANKING_MÊS_AC') or gs('RANKING_MES_AC')
    if ws:
        total_rnums_mi = []
        for row in ws.iter_rows(min_row=1, max_row=40):
            a = row[0].value
            if a and 'TOTAL' in str(a).upper():
                total_rnums_mi.append(row[0].row)
        first_t_mi = total_rnums_mi[0] if total_rnums_mi else 9999
        second_t_mi = total_rnums_mi[1] if len(total_rnums_mi) > 1 else 9999

        ytd_ton_mi = defaultdict(float)
        ytd_fin_mi = defaultdict(float)
        for row in ws.iter_rows(min_row=1, max_row=first_t_mi - 1):
            a, b, c = row[0].value, row[1].value, row[2].value
            if a and str(a).strip().endswith('º') and b and isinstance(b, str):
                ytd_ton_mi[b.strip()] += float(c or 0)
        for row in ws.iter_rows(min_row=first_t_mi + 1, max_row=second_t_mi - 1):
            a, b, c = row[0].value, row[1].value, row[2].value
            if a and str(a).strip().endswith('º') and b and isinstance(b, str):
                ytd_fin_mi[b.strip()] += float(c or 0)

        for name, val in int_u_ton:
            ytd_ton_mi[name] += val
        for name, val in int_u_fin:
            ytd_fin_mi[name] += val

        mi_units_ton = sorted(ytd_ton_mi.items(), key=lambda x: -x[1])
        mi_units_fin = sorted(ytd_fin_mi.items(), key=lambda x: -x[1])
        update_ranking_units(ws, mi_units_ton, mi_units_fin)
        print("  ✓ RANKING_MÊS_AC")

    # 22. UF ranking (interior)
    ws = gs('UF ranking') or gs('UF (2)')
    if ws:
        update_origem_rank(ws, agg_by_uf(all_int_rows))
        print("  ✓ UF ranking")

    # 23. PRODUTOS (interior products)
    ws = gs('PRODUTOS')
    if ws:
        update_rank_sheet(ws, agg_by_produto(all_int_rows))
        print("  ✓ PRODUTOS interior")

    # 24. Update titles across all sheets
    for sn in wb.sheetnames:
        update_titles_in_sheet(wb[sn], mes, ano)
    print("  ✓ Títulos atualizados")

    # 25. Save
    out = io.BytesIO()
    wb.save(out)
    result = out.getvalue()
    print(f"  ✅ Gerado: {len(result):,} bytes")
    return result
